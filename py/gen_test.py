import random
from pathlib import Path

from main import main
from openpyxl import Workbook, load_workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font  # pyright: ignore[reportMissingModuleSource]


ILE_TESTOW = 40
NAZWA_PLIKU = "czasowe.xlsx"

NAZWY_CZASOW = [
    "MCMF [s]",
    "Otoczka [s]",
    "Obwód [s]",
    "SparseTable [s]",
    "Huffman [s]",
    "KMP [s]",
    "Całość [s]"
]


def test(krasnale, miejsce, ataki):
    wyniki = []

    for i in range(ILE_TESTOW):
        liczba_krasnali = random.randint(*krasnale)
        miejsce_w_kopalni = random.randint(*miejsce)
        liczba_atakow = random.randint(*ataki)

        czasy = main(
            liczba_krasnali,
            miejsce_w_kopalni,
            liczba_atakow,
            False,
            False
        )

        wyniki.append((
            i + 1,
            liczba_krasnali,
            miejsce_w_kopalni,
            liczba_atakow,
            czasy
        ))

        print(f"Test {i + 1}: {czasy}")

    if Path(NAZWA_PLIKU).exists():
        workbook = load_workbook(NAZWA_PLIKU)
        arkusz = workbook.active
    else:
        workbook = Workbook()
        arkusz = workbook.active
        arkusz.title = "Wyniki"

    maks_liczba_czasow = max(len(wynik[4]) for wynik in wyniki)

    naglowki_czasow = NAZWY_CZASOW[:maks_liczba_czasow]

    if len(naglowki_czasow) < maks_liczba_czasow:
        for i in range(len(naglowki_czasow), maks_liczba_czasow):
            naglowki_czasow.append(f"Czas {i + 1} [s]")

    arkusz.append([
        "Nr testu",
        "Liczba krasnali",
        "Miejsca w kopalni",
        "Liczba ataków",
        *naglowki_czasow
    ])

    for komorka in arkusz[arkusz.max_row]:
        komorka.font = Font(bold=True)

    for nr_testu, liczba_krasnali, miejsce_w_kopalni, liczba_atakow, czasy in wyniki:
        brakujace = maks_liczba_czasow - len(czasy)

        arkusz.append([
            nr_testu,
            liczba_krasnali,
            miejsce_w_kopalni,
            liczba_atakow,
            *[float(f"{czas:.9f}") for czas in czasy],
            *[""] * brakujace
        ])

    srednie = []

    for indeks_czasu in range(maks_liczba_czasow):
        suma = 0
        ile = 0

        for wynik in wyniki:
            czasy = wynik[4]

            if indeks_czasu < len(czasy):
                suma += czasy[indeks_czasu]
                ile += 1

        srednie.append(suma / ile if ile > 0 else "")

    arkusz.append([])
    arkusz.append([
        "Średni czas",
        "",
        "",
        "",
        *[float(f"{czas:.9f}") if czas != "" else "" for czas in srednie]
    ])

    for komorka in arkusz[arkusz.max_row]:
        komorka.font = Font(bold=True)

    arkusz.append([])

    workbook.save(NAZWA_PLIKU)

    print("Średnie czasy:")
    for nazwa, czas in zip(naglowki_czasow, srednie):
        print(f"{nazwa}: {czas:.9f}")


if __name__ == "__main__":
    krasnale = tuple(map(int, input("Podaj zakres krasnali\n> ").split()))
    miejsce = tuple(map(int, input("Podaj zakres miejsca\n> ").split()))
    ataki = tuple(map(int, input("Podaj zakres ataków\n> ").split()))

    if len(krasnale) == 1:
        krasnale = (*krasnale, *krasnale)

    if len(miejsce) == 1:
        miejsce = (*miejsce, *miejsce)

    if len(ataki) == 1:
        ataki = (*ataki, *ataki)

    test(krasnale, miejsce, ataki)